#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
FX PARITIES ENGINE
- SUPPORTS API MODE OR USER-UPLOADED EXCEL MODE
- BUILDS RAW DAILY FX DATASET WITH FFILL
- BUILDS DASHBOARD-READY MASTERDATA
- EXPORTS FORMATTED EXCEL OUTPUTS
- GENERATES AN INTERACTIVE MULTI-PAGE HTML DASHBOARD
- AUTO-OPENS ALL OUTPUTS
"""

from __future__ import annotations

import json
import math
import os
import platform
import subprocess
import sys
import time
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Dict, Iterable, List, Optional

import pandas as pd
import requests
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

# =========================================================
# CONFIGURATION
# =========================================================
START_DATE = date(2026, 1, 1)
API_BASE = "https://api.frankfurter.dev/v2"
OUTPUT_DIR = Path.cwd() / "FX_PARITIES_OUTPUT"
RAW_SHEET = "FX_PARITIES_DATA"
MASTER_SHEET = "FX_PARITIES_MASTERDATA"
TAB_COLOR = "FF0000"
TITLE_BG = "2D257D"
TITLE_FONT = "FFFFFF"
HEADER_BG = "2D257D"
HEADER_FONT = "FFFFFF"
BODY_BG = "FFF2F2"
THIN_BORDER_COLOR = "D9D9D9"
REQUEST_TIMEOUT = 90
CHUNK_SIZE = 25
SLEEP_BETWEEN_CALLS = 0.08
FOCUS_CURRENCIES = [
    "EUR", "GBP", "SGD", "NOK", "CAD", "MYR", "HKD", "IDR",
    "NZD", "CNY", "INR", "BRL", "AUD", "THB", "ZAR"
]
OTHER_PAGE_SIZE = 12

RAW_COLUMNS = [
    "DATE",
    "CURRENCY_CODE",
    "USD_RATE",
    "PROVIDER",
    "IS_FILLED",
    "LOAD_TIMESTAMP",
]

MASTER_COLUMNS = [
    "DATE",
    "CURRENCY_CODE",
    "CURRENCY_NAME",
    "USD_RATE",
    "PROVIDER",
    "IS_FILLED",
    "LOAD_TIMESTAMP",
    "YEAR",
    "MONTH_NUM",
    "MONTH_LABEL",
    "MONTH_NAME",
    "WEEKDAY_NAME",
    "IS_WEEKEND",
    "RATE_CHANGE_DOD_PCT",
    "RATE_CHANGE_MTD_PCT",
    "RATE_CHANGE_YTD_PCT",
    "RATE_INDEX_100",
    "PRIOR_MONTH_AVG_RATE",
    "CHANGE_VS_PRIOR_MONTH_AVG_PCT",
    "ROLLING_30D_VOLATILITY_PCT",
]

STATIC_CURRENCY_NAMES = {
    "USD": "US DOLLAR",
    "EUR": "EURO",
    "GBP": "POUND STERLING",
    "SGD": "SINGAPORE DOLLAR",
    "NOK": "NORWEGIAN KRONE",
    "CAD": "CANADIAN DOLLAR",
    "MYR": "MALAYSIAN RINGGIT",
    "HKD": "HONG KONG DOLLAR",
    "IDR": "INDONESIAN RUPIAH",
    "NZD": "NEW ZEALAND DOLLAR",
    "CNY": "CHINESE YUAN",
    "INR": "INDIAN RUPEE",
    "BRL": "BRAZILIAN REAL",
    "AUD": "AUSTRALIAN DOLLAR",
    "THB": "THAI BAHT",
    "AOA": "ANGOLAN KWANZA",
    "ZAR": "SOUTH AFRICAN RAND",
    "JPY": "JAPANESE YEN",
    "CHF": "SWISS FRANC",
    "SEK": "SWEDISH KRONA",
    "DKK": "DANISH KRONE",
    "PLN": "POLISH ZLOTY",
    "CZK": "CZECH KORUNA",
    "HUF": "HUNGARIAN FORINT",
    "RON": "ROMANIAN LEU",
    "TRY": "TURKISH LIRA",
    "AED": "UAE DIRHAM",
    "SAR": "SAUDI RIYAL",
    "QAR": "QATARI RIYAL",
    "BHD": "BAHRAINI DINAR",
    "KWD": "KUWAITI DINAR",
    "OMR": "OMANI RIAL",
    "KRW": "SOUTH KOREAN WON",
    "TWD": "NEW TAIWAN DOLLAR",
    "PHP": "PHILIPPINE PESO",
    "VND": "VIETNAMESE DONG",
    "MXN": "MEXICAN PESO",
    "CLP": "CHILEAN PESO",
    "COP": "COLOMBIAN PESO",
    "PEN": "PERUVIAN SOL",
    "ARS": "ARGENTINE PESO",
    "ILS": "ISRAELI SHEKEL",
    "EGP": "EGYPTIAN POUND",
    "KES": "KENYAN SHILLING",
    "NGN": "NIGERIAN NAIRA",
    "GHS": "GHANAIAN CEDI",
    "MAD": "MOROCCAN DIRHAM",
    "RUB": "RUSSIAN RUBLE",
    "UAH": "UKRAINIAN HRYVNIA",
}

YES_VALUES = {"Y", "YES", "YE", "YEAH", "YEP", "TRUE", "1"}
NO_VALUES = {"N", "NO", "NOPE", "FALSE", "0"}


# =========================================================
# TERMINAL HELPERS
# =========================================================
def info(message: str) -> None:
    print(f"[INFO] {str(message).upper()}")


def warn(message: str) -> None:
    print(f"[WARN] {str(message).upper()}")


def done(message: str) -> None:
    print(f"[DONE] {str(message).upper()}")


def fail(message: str) -> None:
    print(f"[ERROR] {str(message).upper()}")


def local_today() -> date:
    return datetime.now().astimezone().date()


def local_timestamp() -> str:
    return datetime.now().astimezone().strftime("%Y-%m-%d %H:%M:%S")


def build_output_paths(today: date) -> Dict[str, Path]:
    stamp = today.strftime("%Y%m%d")
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    return {
        "rates_xlsx": OUTPUT_DIR / f"FX_PARITIES_RATES_{stamp}.xlsx",
        "master_xlsx": OUTPUT_DIR / f"FX_PARITIES_MASTERDATA_{stamp}.xlsx",
        "dashboard_html": OUTPUT_DIR / f"FX_PARITIES_DASHBOARD_{stamp}.html",
    }


def print_banner(report_date: date) -> None:
    line = "=" * 120
    print(line)
    print("FX PARITIES ENGINE")
    print(line)
    print("THIS ENGINE CAN GENERATE AN FX REPORT IN TWO DIFFERENT MODES:")
    print("1) API MODE: DAILY USD-BASE FX RATES WILL BE FETCHED AUTOMATICALLY.")
    print("2) USER DATA MODE: YOUR OWN EXCEL FILE CAN BE LOADED AND RESHAPED INTO THE SAME OUTPUT STRUCTURE.")
    print("THE ENGINE WILL CREATE A RAW FX DATA WORKBOOK, A STANDALONE MASTERDATA WORKBOOK, AND AN")
    print("INTERACTIVE MULTI-PAGE HTML DASHBOARD, AND THEN IT WILL OPEN ALL OUTPUTS AUTOMATICALLY.")
    print(f"DEFAULT API START DATE: {START_DATE.isoformat()}")
    print(f"LOCAL REPORT DATE: {report_date.isoformat()}")
    print(line)
    input(f"PLEASE PRESS START TO GENERATE FX REPORT FOR {report_date.year} {report_date.month:02d} {report_date.day:02d}")


def finish_banner(paths: List[Path]) -> None:
    line = "=" * 120
    print(line)
    print("PROCESS COMPLETED SUCCESSFULLY.")
    print("THE FOLLOWING OUTPUTS WERE GENERATED:")
    for p in paths:
        print(f" - {str(p)}")
    print(f"ALL OUTPUTS WERE SAVED UNDER: {OUTPUT_DIR}")
    print(line)
    input("PRESS ENTER TO FINISH THE PROCESS")


def ask_yes_no(prompt: str) -> bool:
    while True:
        answer = input(prompt + " ").strip().upper()
        if answer in YES_VALUES:
            return True
        if answer in NO_VALUES:
            return False
        warn("INVALID ANSWER. PLEASE TYPE YES OR NO.")


def ask_existing_file_path(prompt: str) -> Path:
    while True:
        raw = input(prompt + " ").strip().strip('"').strip("'")
        path = Path(raw)
        if path.exists() and path.is_file():
            return path
        warn("THE FILE PATH PROVIDED COULD NOT BE FOUND. PLEASE TRY AGAIN.")


# =========================================================
# API HELPERS
# =========================================================
def session_with_retries() -> requests.Session:
    session = requests.Session()
    session.headers.update(
        {
            "User-Agent": "FX-PARITIES-ENGINE/2.0",
            "Accept": "application/json",
        }
    )
    return session


def get_json(session: requests.Session, url: str, params: Optional[dict] = None):
    last_error = None
    for attempt in range(1, 4):
        try:
            response = session.get(url, params=params, timeout=REQUEST_TIMEOUT)
            response.raise_for_status()
            return response.json()
        except Exception as exc:  # noqa: BLE001
            last_error = exc
            warn(f"REQUEST FAILED ON ATTEMPT {attempt}: {exc}")
            time.sleep(1.2 * attempt)
    raise RuntimeError(f"API REQUEST FAILED AFTER 3 ATTEMPTS: {last_error}")


def normalize_currencies_payload(payload) -> pd.DataFrame:
    if isinstance(payload, list):
        df = pd.DataFrame(payload)
        if "iso_code" not in df.columns:
            raise RuntimeError("THE CURRENCIES RESPONSE DID NOT CONTAIN ISO_CODE.")
        df["name"] = df.get("name", df["iso_code"])
        return df[["iso_code", "name"]].drop_duplicates().reset_index(drop=True)

    if isinstance(payload, dict):
        # v1-style fallback: {"EUR": "Euro", ...}
        rows = [{"iso_code": k, "name": v} for k, v in payload.items()]
        return pd.DataFrame(rows)

    raise RuntimeError("UNSUPPORTED CURRENCIES RESPONSE FORMAT.")


def fetch_active_currencies(session: requests.Session, end_date: date) -> pd.DataFrame:
    info("FETCHING AVAILABLE CURRENCIES FROM FRANKFURTER")
    payload = get_json(session, f"{API_BASE}/currencies")
    df = normalize_currencies_payload(payload)
    if df.empty:
        raise RuntimeError("NO CURRENCY METADATA RETURNED FROM THE API.")

    df["iso_code"] = df["iso_code"].astype(str).str.upper().str.strip()
    df["name"] = df["name"].astype(str).str.upper().str.strip()

    if not (df["iso_code"] == "USD").any():
        usd_row = pd.DataFrame([{"iso_code": "USD", "name": "US DOLLAR"}])
        df = pd.concat([usd_row, df], ignore_index=True)

    df = df.sort_values("iso_code").drop_duplicates(subset=["iso_code"]).reset_index(drop=True)
    done(f"ACTIVE CURRENCIES IDENTIFIED: {len(df):,}")
    return df


def chunked(items: List[str], size: int) -> Iterable[List[str]]:
    for i in range(0, len(items), size):
        yield items[i:i + size]


def fetch_rate_chunk(
    session: requests.Session,
    quotes: List[str],
    start_date: date,
    end_date: date,
) -> pd.DataFrame:
    if not quotes:
        return pd.DataFrame(columns=["date", "currency_code", "usd_rate"])

    params = {
        "base": "USD",
        "from": start_date.isoformat(),
        "to": end_date.isoformat(),
        "quotes": ",".join(quotes),
    }
    payload = get_json(session, f"{API_BASE}/rates", params=params)
    df = pd.DataFrame(payload)
    if df.empty:
        return pd.DataFrame(columns=["date", "currency_code", "usd_rate"])

    expected = {"date", "quote", "rate"}
    if not expected.issubset(df.columns):
        raise RuntimeError(f"RATES RESPONSE IS MISSING EXPECTED FIELDS: {sorted(expected.difference(df.columns))}")

    df = df[["date", "quote", "rate"]].copy()
    df.rename(columns={"quote": "currency_code", "rate": "usd_rate"}, inplace=True)
    return df


def fetch_observed_rates_api(
    session: requests.Session,
    currency_codes: List[str],
    start_date: date,
    end_date: date,
    load_ts: str,
) -> pd.DataFrame:
    info("FETCHING OBSERVED FX RATES FROM THE API IN QUOTED CHUNKS")
    frames: List[pd.DataFrame] = []

    usd_frame = pd.DataFrame(
        {
            "date": pd.date_range(start_date, end_date, freq="D").date,
            "currency_code": "USD",
            "usd_rate": 1.0,
        }
    )
    frames.append(usd_frame)

    non_usd = [c for c in currency_codes if c != "USD"]
    chunks = list(chunked(non_usd, CHUNK_SIZE))

    for idx, chunk in enumerate(chunks, start=1):
        info(f"FETCHING API CHUNK {idx} OF {len(chunks)} WITH {len(chunk)} CURRENCIES")
        df = fetch_rate_chunk(session, chunk, start_date, end_date)
        if not df.empty:
            frames.append(df)
        time.sleep(SLEEP_BETWEEN_CALLS)

    observed = pd.concat(frames, ignore_index=True)
    observed["date"] = pd.to_datetime(observed["date"], errors="coerce").dt.date
    observed["currency_code"] = observed["currency_code"].astype(str).str.upper().str.strip()
    observed["usd_rate"] = pd.to_numeric(observed["usd_rate"], errors="coerce")
    observed = observed.dropna(subset=["date", "currency_code", "usd_rate"])
    observed["provider"] = "FRANKFURTER"
    observed["is_filled"] = False
    observed["load_timestamp"] = load_ts
    observed = observed.drop_duplicates(subset=["date", "currency_code"]).sort_values(["currency_code", "date"])
    done(f"OBSERVED API RATE ROWS COLLECTED: {len(observed):,}")
    return observed.reset_index(drop=True)


# =========================================================
# USER-UPLOAD HELPERS
# =========================================================
def normalize_columns(columns: Iterable[str]) -> List[str]:
    normalized = []
    for col in columns:
        text = str(col).strip().upper().replace(" ", "_")
        normalized.append(text)
    return normalized


def normalize_bool_series(series: pd.Series) -> pd.Series:
    def _conv(v):
        if pd.isna(v):
            return False
        s = str(v).strip().upper()
        if s in {"TRUE", "T", "YES", "Y", "1"}:
            return True
        if s in {"FALSE", "F", "NO", "N", "0"}:
            return False
        return False

    return series.apply(_conv)


def load_user_excel_raw(file_path: Path, load_ts: str) -> pd.DataFrame:
    info("LOADING USER-PROVIDED EXCEL FILE FROM THE FIRST SHEET")
    df = pd.read_excel(file_path, sheet_name=0)
    if df.empty:
        raise RuntimeError("THE FIRST SHEET OF THE PROVIDED EXCEL FILE IS EMPTY.")

    df.columns = normalize_columns(df.columns)

    required = {"DATE", "CURRENCY_CODE", "USD_RATE"}
    if not required.issubset(df.columns):
        raise RuntimeError(
            f"THE PROVIDED FILE DOES NOT CONTAIN THE REQUIRED COLUMNS: {sorted(required)}"
        )

    for col in ["PROVIDER", "IS_FILLED", "LOAD_TIMESTAMP"]:
        if col not in df.columns:
            df[col] = None

    df = df[["DATE", "CURRENCY_CODE", "USD_RATE", "PROVIDER", "IS_FILLED", "LOAD_TIMESTAMP"]].copy()
    df["DATE"] = pd.to_datetime(df["DATE"], errors="coerce").dt.date
    df["CURRENCY_CODE"] = df["CURRENCY_CODE"].astype(str).str.upper().str.strip()
    df["USD_RATE"] = pd.to_numeric(df["USD_RATE"], errors="coerce")
    df["PROVIDER"] = df["PROVIDER"].fillna("USER_UPLOAD").astype(str).str.upper().str.strip()
    df["IS_FILLED"] = normalize_bool_series(df["IS_FILLED"])
    df["LOAD_TIMESTAMP"] = df["LOAD_TIMESTAMP"].fillna(load_ts).astype(str)

    df = df.dropna(subset=["DATE", "CURRENCY_CODE", "USD_RATE"])
    df = df.drop_duplicates(subset=["DATE", "CURRENCY_CODE"], keep="last")
    df = df.sort_values(["CURRENCY_CODE", "DATE"]).reset_index(drop=True)
    done(f"USER DATA ROWS LOADED: {len(df):,}")
    return df


def build_currency_catalog_from_upload(df: pd.DataFrame) -> pd.DataFrame:
    codes = sorted(df["CURRENCY_CODE"].dropna().astype(str).str.upper().unique().tolist())
    rows = []
    for code in codes:
        rows.append({"iso_code": code, "name": STATIC_CURRENCY_NAMES.get(code, code)})
    if "USD" not in codes:
        rows.append({"iso_code": "USD", "name": STATIC_CURRENCY_NAMES.get("USD", "USD")})
    out = pd.DataFrame(rows).drop_duplicates(subset=["iso_code"]).sort_values("iso_code").reset_index(drop=True)
    return out


# =========================================================
# SHAPING HELPERS
# =========================================================
def build_full_calendar_dataset(
    observed: pd.DataFrame,
    currencies_df: pd.DataFrame,
    start_date: date,
    end_date: date,
    default_provider: str,
    load_ts: str,
) -> pd.DataFrame:
    info("BUILDING FULL DAILY CALENDAR AND APPLYING FFILL FOR MISSING DAYS")
    all_dates = pd.date_range(start_date, end_date, freq="D").date
    result_frames: List[pd.DataFrame] = []
    active_codes = sorted(currencies_df["iso_code"].astype(str).str.upper().tolist())

    for code in active_codes:
        sub = observed.loc[observed["currency_code"] == code].copy()
        if sub.empty:
            warn(f"NO OBSERVED DATA FOUND FOR {code}; THIS CURRENCY WILL BE SKIPPED")
            continue

        base = pd.DataFrame({"date": all_dates})
        merged = base.merge(
            sub[["date", "usd_rate", "provider"]],
            on="date",
            how="left",
        )
        observed_mask = merged["usd_rate"].notna()
        merged["usd_rate"] = pd.to_numeric(merged["usd_rate"], errors="coerce").ffill().bfill()
        if merged["usd_rate"].isna().all():
            warn(f"{code} COULD NOT BE FILLED AND WILL BE SKIPPED")
            continue

        merged["provider"] = merged["provider"].ffill().bfill().fillna(default_provider)
        merged["currency_code"] = code
        merged["is_filled"] = ~observed_mask
        merged["load_timestamp"] = load_ts
        result_frames.append(merged[["date", "currency_code", "usd_rate", "provider", "is_filled", "load_timestamp"]])

    if not result_frames:
        raise RuntimeError("NO FX DATA COULD BE PREPARED AFTER THE FFILL STEP.")

    raw_df = pd.concat(result_frames, ignore_index=True)
    raw_df["date"] = pd.to_datetime(raw_df["date"], errors="coerce")
    raw_df["currency_code"] = raw_df["currency_code"].astype(str).str.upper().str.strip()
    raw_df["provider"] = raw_df["provider"].astype(str).str.upper().str.strip()
    raw_df["usd_rate"] = pd.to_numeric(raw_df["usd_rate"], errors="coerce")
    raw_df = raw_df.dropna(subset=["date", "currency_code", "usd_rate"])
    raw_df = raw_df.drop_duplicates(subset=["date", "currency_code"], keep="last")
    raw_df = raw_df.sort_values(["date", "currency_code"]).reset_index(drop=True)
    raw_df["date"] = raw_df["date"].dt.date
    done(f"FINAL RAW DATASET ROWS AFTER FFILL: {len(raw_df):,}")
    return raw_df


def build_masterdata(raw_df: pd.DataFrame, currencies_df: pd.DataFrame) -> pd.DataFrame:
    info("BUILDING DASHBOARD-READY MASTERDATA")
    name_map = currencies_df[["iso_code", "name"]].copy()
    name_map["iso_code"] = name_map["iso_code"].astype(str).str.upper().str.strip()
    name_map["name"] = name_map["name"].astype(str).str.upper().str.strip()
    name_map.rename(columns={"iso_code": "currency_code", "name": "currency_name"}, inplace=True)

    df = raw_df.copy()
    df = df.merge(name_map, on="currency_code", how="left")
    df["currency_name"] = df["currency_name"].fillna(df["currency_code"])
    df["date"] = pd.to_datetime(df["date"])
    df = df.sort_values(["currency_code", "date"]).reset_index(drop=True)

    df["year"] = df["date"].dt.year
    df["month_num"] = df["date"].dt.month
    df["month_label"] = df["date"].dt.strftime("%Y-%m")
    df["month_name"] = df["date"].dt.strftime("%B").str.upper()
    df["weekday_name"] = df["date"].dt.strftime("%A").str.upper()
    df["is_weekend"] = df["date"].dt.weekday >= 5

    df["rate_change_dod_pct"] = df.groupby("currency_code")["usd_rate"].pct_change()
    first_rate = df.groupby("currency_code")["usd_rate"].transform("first")
    month_first = df.groupby(["currency_code", "month_label"])["usd_rate"].transform("first")

    df["rate_change_ytd_pct"] = (df["usd_rate"] / first_rate) - 1.0
    df["rate_change_mtd_pct"] = (df["usd_rate"] / month_first) - 1.0
    df["rate_index_100"] = (df["usd_rate"] / first_rate) * 100.0

    month_avg = (
        df.groupby(["currency_code", "month_label"], as_index=False)["usd_rate"]
        .mean()
        .rename(columns={"usd_rate": "month_avg_rate"})
    )
    month_avg["month_period"] = pd.PeriodIndex(month_avg["month_label"], freq="M")
    prev_month_avg = month_avg.copy()
    prev_month_avg["month_period"] = prev_month_avg["month_period"] + 1
    prev_month_avg.rename(columns={"month_avg_rate": "prior_month_avg_rate"}, inplace=True)

    df["month_period"] = pd.PeriodIndex(df["month_label"], freq="M")
    df = df.merge(
        prev_month_avg[["currency_code", "month_period", "prior_month_avg_rate"]],
        on=["currency_code", "month_period"],
        how="left",
    )
    df["change_vs_prior_month_avg_pct"] = (df["usd_rate"] / df["prior_month_avg_rate"]) - 1.0
    df.loc[df["prior_month_avg_rate"].isna(), "change_vs_prior_month_avg_pct"] = pd.NA

    vol = (
        df.groupby("currency_code")["rate_change_dod_pct"]
        .rolling(30, min_periods=5)
        .std()
        .reset_index(level=0, drop=True)
    )
    df["rolling_30d_volatility_pct"] = vol

    keep = [
        "date",
        "currency_code",
        "currency_name",
        "usd_rate",
        "provider",
        "is_filled",
        "load_timestamp",
        "year",
        "month_num",
        "month_label",
        "month_name",
        "weekday_name",
        "is_weekend",
        "rate_change_dod_pct",
        "rate_change_mtd_pct",
        "rate_change_ytd_pct",
        "rate_index_100",
        "prior_month_avg_rate",
        "change_vs_prior_month_avg_pct",
        "rolling_30d_volatility_pct",
    ]
    df = df[keep].copy()
    done(f"MASTERDATA ROWS PREPARED: {len(df):,}")
    return df


# =========================================================
# EXCEL HELPERS
# =========================================================
def to_excel_value(value):
    if pd.isna(value):
        return None
    if isinstance(value, pd.Timestamp):
        return value.to_pydatetime()
    return value


def auto_width(ws) -> None:
    max_widths: Dict[int, int] = {}
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            text = str(cell.value)
            max_widths[cell.column] = max(max_widths.get(cell.column, 0), len(text))
    for col_idx, length in max_widths.items():
        width = min(max(length + 3, 12), 42)
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def style_sheet(ws, df: pd.DataFrame, subtitle: str) -> None:
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = TAB_COLOR
    ws.freeze_panes = "A5"

    max_col = len(df.columns)
    title_fill = PatternFill("solid", fgColor=TITLE_BG)
    header_fill = PatternFill("solid", fgColor=HEADER_BG)
    body_fill = PatternFill("solid", fgColor=BODY_BG)
    thin = Side(style="thin", color=THIN_BORDER_COLOR)
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_col)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=max_col)
    ws["A1"] = "FOREIGN EXCHANGE PARITIES - 1 USD EQUALS TO"
    ws["A2"] = subtitle.upper()

    for coord in ("A1", "A2"):
        c = ws[coord]
        c.fill = title_fill
        c.font = Font(bold=True, color=TITLE_FONT, size=12)
        c.alignment = Alignment(horizontal="center", vertical="center")

    ws.row_dimensions[1].height = 24
    ws.row_dimensions[2].height = 21

    header_row = 4
    for cell in ws[header_row]:
        if cell.column <= max_col:
            cell.fill = header_fill
            cell.font = Font(bold=True, color=HEADER_FONT)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border

    for row in ws.iter_rows(min_row=5, max_row=ws.max_row, min_col=1, max_col=max_col):
        for cell in row:
            cell.font = Font(bold=True, color="000000")
            cell.fill = body_fill
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center")

    for col_idx, name in enumerate(df.columns, start=1):
        col_letter = get_column_letter(col_idx)
        if name == "DATE":
            for cell in ws[col_letter][4:]:
                cell.number_format = "yyyy-mm-dd"
        elif name.endswith("_PCT"):
            for cell in ws[col_letter][4:]:
                cell.number_format = "0.00%"
        elif name in {"USD_RATE", "RATE_INDEX_100", "PRIOR_MONTH_AVG_RATE"}:
            for cell in ws[col_letter][4:]:
                cell.number_format = "0.000000"
        elif name in {"YEAR", "MONTH_NUM"}:
            for cell in ws[col_letter][4:]:
                cell.number_format = "0"

    auto_width(ws)


def add_dataframe_as_table(ws, df: pd.DataFrame, table_name: str) -> None:
    start_row = 4
    start_col = 1

    for col_idx, col_name in enumerate(df.columns, start=start_col):
        ws.cell(row=start_row, column=col_idx, value=col_name.upper())

    for row_idx, row in enumerate(df.itertuples(index=False), start=start_row + 1):
        for col_idx, value in enumerate(row, start=start_col):
            cell = ws.cell(row=row_idx, column=col_idx, value=to_excel_value(value))
            if isinstance(value, str):
                cell.value = value.upper()

    last_row = start_row + len(df)
    last_col = start_col + len(df.columns) - 1
    ref = f"A{start_row}:{get_column_letter(last_col)}{last_row}"
    table = Table(displayName=table_name, ref=ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=False,
        showColumnStripes=False,
    )
    ws.add_table(table)


def prepare_output_dataframe(df: pd.DataFrame, columns: List[str]) -> pd.DataFrame:
    out = df.copy()
    out.columns = [c.upper() for c in out.columns]
    out = out[columns].copy()
    for col in out.columns:
        if out[col].dtype == "bool":
            out[col] = out[col].map({True: "TRUE", False: "FALSE"})
    for col in out.select_dtypes(include=["object"]).columns:
        out[col] = out[col].astype(str).str.upper()
    if "DATE" in out.columns:
        out["DATE"] = pd.to_datetime(out["DATE"]).dt.date
    return out


def write_workbook(path: Path, raw_df: pd.DataFrame, master_df: pd.DataFrame) -> None:
    wb = Workbook()
    wb.remove(wb.active)

    raw_ws = wb.create_sheet(RAW_SHEET)
    master_ws = wb.create_sheet(MASTER_SHEET)

    raw_out = prepare_output_dataframe(raw_df, RAW_COLUMNS)
    master_out = prepare_output_dataframe(master_df, MASTER_COLUMNS)

    add_dataframe_as_table(raw_ws, raw_out, "TBL_FX_PARITIES_DATA")
    add_dataframe_as_table(master_ws, master_out, "TBL_FX_PARITIES_MASTERDATA")

    style_sheet(raw_ws, raw_out, f"{RAW_SHEET} | DAILY RAW FX DATASET WITH FFILL")
    style_sheet(master_ws, master_out, f"{MASTER_SHEET} | DASHBOARD READY MASTERDATA")
    wb.save(path)


def write_master_only_workbook(path: Path, master_df: pd.DataFrame) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = MASTER_SHEET
    master_out = prepare_output_dataframe(master_df, MASTER_COLUMNS)
    add_dataframe_as_table(ws, master_out, "TBL_FX_PARITIES_MASTERDATA_EXPORT")
    style_sheet(ws, master_out, f"{MASTER_SHEET} | STANDALONE EXPORT")
    wb.save(path)


# =========================================================
# DASHBOARD PAYLOAD / HTML
# =========================================================
def build_dashboard_payload(master_df: pd.DataFrame) -> dict:
    df = master_df.copy()
    df["date"] = pd.to_datetime(df["date"])
    latest_date = df["date"].max()

    payload_df = df.copy()
    payload_df["date"] = payload_df["date"].dt.strftime("%Y-%m-%d")
    payload_df = payload_df.replace({pd.NA: None, math.inf: None, -math.inf: None})

    return {
        "meta": {
            "title": "FX PARITIES DASHBOARD",
            "start_date": df["date"].min().strftime("%Y-%m-%d"),
            "end_date": latest_date.strftime("%Y-%m-%d"),
            "generated_at": local_timestamp(),
            "currencies_tracked": int(df["currency_code"].nunique()),
            "calendar_days": int(df["date"].nunique()),
            "focus_currencies": FOCUS_CURRENCIES,
            "other_page_size": OTHER_PAGE_SIZE,
        },
        "records": json.loads(payload_df.to_json(orient="records")),
    }



def dashboard_html(payload: dict) -> str:
    template = r'''<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>FX PARITIES DASHBOARD</title>
<link href="https://fonts.googleapis.com/css2?family=DM+Serif+Display:ital@0;1&family=DM+Mono:wght@400;500&family=DM+Sans:opsz,wght@9..40,300;9..40,400;9..40,500&display=swap" rel="stylesheet">
<script src="https://cdnjs.cloudflare.com/ajax/libs/Chart.js/4.4.1/chart.umd.js"></script>
<style>
:root{
  --bg:#07080b;--bg2:#10141d;--bg3:#151b26;--bg4:#1d2433;
  --border:rgba(255,255,255,.14);--border2:rgba(255,255,255,.28);--border3:rgba(255,255,255,.48);
  --text:#ffffff;--text2:#e7e2dc;--text3:#b9b1a8;--text4:#7b736b;
  --gold:#e2b347;--gold2:#f2c96a;--gold3:#ffe3a0;
  --green:#34d27c;--red:#ff6b6b;--blue:#66b3ff;--purple:#bb7cff;--teal:#35d1c3;--orange:#ff9f43;
}
*{box-sizing:border-box;margin:0;padding:0}
html{scroll-behavior:smooth}
body{font-family:'DM Sans',sans-serif;background:var(--bg);color:var(--text);min-height:100vh}
button,select,input{font-family:inherit}
select option{background:#0f141d;color:#ffffff}
input[type="date"]::-webkit-calendar-picker-indicator{filter:invert(1) brightness(1.8);opacity:1;cursor:pointer}
header{position:sticky;top:0;z-index:200;background:rgba(7,8,11,.97);backdrop-filter:blur(16px);border-bottom:2px solid var(--border2);display:grid;grid-template-columns:1fr auto 1.15fr;align-items:center;padding:0 1.6rem;min-height:124px;gap:16px}
.logo{display:flex;align-items:center;gap:14px}.logo-mark{width:40px;height:40px;border-radius:10px;background:linear-gradient(135deg,var(--gold),#a07830);display:flex;align-items:center;justify-content:center;flex-shrink:0;box-shadow:0 0 0 2px rgba(255,255,255,.18)}
.logo-mark svg{width:20px;height:20px;stroke:#07080b;fill:none;stroke-width:2;stroke-linecap:round;stroke-linejoin:round}.logo-title{font-family:'DM Serif Display',serif;font-size:21px;color:var(--gold2);letter-spacing:.01em;font-weight:400}.logo-sub{font-size:12px;font-weight:700;color:var(--gold2);letter-spacing:.06em;text-transform:uppercase;margin-top:3px}
.header-center{text-align:center;display:flex;flex-direction:column;align-items:center;justify-content:center;padding:8px 0 10px}.header-center-title{font-family:'DM Serif Display',serif;font-size:20px;color:#ffffff;font-weight:400;letter-spacing:.02em;line-height:1.15}.header-center-sub{font-size:12px;font-weight:800;color:var(--gold2);letter-spacing:.08em;text-transform:uppercase;margin-top:6px;line-height:1.2}
.header-right{display:flex;flex-direction:column;align-items:flex-end;gap:8px}.header-stats{display:flex;flex-direction:column;align-items:flex-end;gap:4px}.clock{font-family:'DM Mono',monospace;font-size:12px;font-weight:700;color:#ffffff;letter-spacing:.06em}.dyn-range{font-family:'DM Mono',monospace;font-size:11px;font-weight:800;color:var(--gold2);letter-spacing:.07em;text-transform:uppercase}
.insight-box{width:min(560px,100%);padding:.72rem .9rem;border:2px solid rgba(255,255,255,.75);border-radius:10px;background:linear-gradient(180deg,rgba(22,28,39,.96),rgba(10,13,19,.96));box-shadow:0 0 0 1px rgba(255,255,255,.05) inset}.insight-box-label{font-size:10px;font-weight:800;letter-spacing:.11em;text-transform:uppercase;color:var(--gold2);margin-bottom:.42rem}.insight-box-text{font-size:11px;line-height:1.55;color:#ffffff}
.toolbar{display:flex;align-items:center;gap:10px;padding:.85rem 1.4rem .75rem;border-bottom:2px solid var(--border);flex-wrap:wrap}.tool-chip{display:flex;align-items:center;gap:10px;padding:8px 12px;border:2px solid rgba(255,255,255,.7);border-radius:999px;background:rgba(15,20,29,.95);box-shadow:0 0 0 1px rgba(255,255,255,.03) inset}.tool-chip label{font-family:'DM Mono',monospace;font-size:10px;font-weight:800;color:var(--text2);letter-spacing:.08em;text-transform:uppercase}.tool-chip select,.tool-chip input{background:transparent;border:none;color:#ffffff;font-family:'DM Mono',monospace;font-size:11px;font-weight:700;outline:none}.tool-chip input[type="date"]{min-width:128px}
.clear-btn{border:2px solid rgba(255,255,255,.75);background:transparent;color:#ffffff;cursor:pointer;font-family:'DM Mono',monospace;font-size:12px;font-weight:800;line-height:1;border-radius:999px;width:22px;height:22px;display:flex;align-items:center;justify-content:center}.clear-btn:hover{background:rgba(255,255,255,.08)}
.view-bar{display:flex;align-items:center;gap:8px;padding:0 1.4rem .8rem;border-bottom:1px solid var(--border)}
.view-btn,.page-btn{padding:7px 14px;font-size:11px;font-weight:800;font-family:'DM Mono',monospace;letter-spacing:.06em;border:2px solid rgba(255,255,255,.78);border-radius:999px;background:transparent;color:#ffffff;cursor:pointer;transition:all .15s}.view-btn:hover,.page-btn:hover{border-color:var(--gold2);color:var(--gold2)}.view-btn.on,.page-btn.on{background:var(--gold);color:#07080b;border-color:#ffffff}
.page-bar{display:flex;gap:8px;align-items:center;padding:.75rem 1.4rem;border-bottom:2px solid var(--border);flex-wrap:wrap}.active-filter{padding:0 1.4rem 1rem;color:var(--text3);font-size:11px;font-family:'DM Mono',monospace;letter-spacing:.04em;text-transform:uppercase}
.page-shell{display:flex;flex-direction:column;gap:1px;background:var(--border2)}
.kpi-strip{display:grid;grid-template-columns:repeat(3,1fr);gap:1px;background:var(--border2)}.kpi{background:var(--bg2);padding:1.38rem 1.2rem 1.28rem;position:relative;overflow:hidden;display:flex;flex-direction:column;justify-content:center;align-items:center;text-align:center;min-height:132px;border:2px solid rgba(255,255,255,.22);box-shadow:0 0 28px rgba(0,0,0,.18) inset}.kpi::after{content:'';position:absolute;top:0;left:0;right:0;height:3px;background:linear-gradient(90deg,transparent,var(--kc,var(--gold)),transparent);opacity:.95}.kpi-label{font-size:11px;font-weight:900;letter-spacing:.12em;text-transform:uppercase;color:var(--text2);margin-bottom:.72rem;line-height:1.25}.kpi-val{font-family:'DM Serif Display',serif;font-size:31px;line-height:1.02;margin-bottom:.42rem;color:#ffffff}.kpi-sub{font-family:'DM Mono',monospace;font-size:10px;color:var(--text3);line-height:1.4;max-width:100%}
.card-grid{display:grid;grid-template-columns:repeat(auto-fit,minmax(148px,1fr));gap:1px;background:var(--border2)}.rate-card{background:var(--bg2);padding:1rem .95rem .92rem;min-height:112px;border:2px solid rgba(255,255,255,.22);box-shadow:0 0 28px rgba(0,0,0,.15) inset}.rate-card-top{display:flex;align-items:flex-start;justify-content:space-between;gap:8px;margin-bottom:.42rem}.rate-card-code{font-size:13px;font-weight:900;color:#ffffff;letter-spacing:.08em;text-transform:uppercase}.rate-card-name{font-size:10px;color:var(--text3);letter-spacing:.03em;text-transform:uppercase;line-height:1.25;max-width:70%}.rate-card-rate{font-family:'DM Serif Display',serif;font-size:24px;color:var(--gold3);line-height:1.05;margin-bottom:.36rem}.rate-card-delta{font-family:'DM Mono',monospace;font-size:10px;font-weight:800}.rate-card-delta.pos{color:var(--green)}.rate-card-delta.neg{color:var(--red)}
.section{background:var(--bg2);padding:1.12rem 1.2rem;border:2px solid rgba(255,255,255,.18)}.section-title{font-size:14px;font-weight:800;color:#ffffff;margin-bottom:.95rem;display:flex;align-items:center;gap:10px;background:#2D257D;padding:.58rem .95rem;border-radius:4px;box-shadow:0 0 0 1px rgba(255,255,255,.08) inset}.section-title::before{content:'';width:3px;height:14px;background:var(--gold2);border-radius:2px;flex-shrink:0}.section-note{padding:0 0 .35rem;color:var(--text3);font-size:10px;font-family:'DM Mono',monospace;letter-spacing:.06em;text-transform:uppercase}
.chart-grid{display:grid;grid-template-columns:repeat(var(--cols,3),minmax(0,1fr));gap:1px;background:var(--border2);border-top:2px solid var(--border)}.chart-card{background:var(--bg2);padding:1rem 1rem 1.1rem;display:flex;flex-direction:column;min-height:370px;border:2px solid rgba(255,255,255,.22)}.chart-card[data-span="full"]{grid-column:1 / -1}.chart-card[data-span="half"]{grid-column:span 2}.chart-head{display:flex;align-items:center;justify-content:space-between;gap:10px;margin-bottom:.72rem}.chart-title{font-size:12px;font-weight:800;color:#ffffff;background:#2D257D;padding:.36rem .72rem;border-radius:3px;display:inline-flex;align-items:center;gap:8px}.chart-title::before{content:'';width:2px;height:12px;background:var(--gold2);border-radius:2px;display:inline-block}.chart-sub{font-size:10px;color:var(--text3);font-family:'DM Mono',monospace;text-transform:uppercase}.chart-wrap{position:relative;width:100%;height:285px;flex:1}
.insight-grid{display:grid;grid-template-columns:1.1fr .95fr .95fr;gap:1px;background:var(--border2);border-top:2px solid var(--border)}.panel{background:var(--bg2);padding:1.12rem 1.18rem;border:2px solid rgba(255,255,255,.2)}.panel-title{font-size:14px;font-weight:800;color:#ffffff;margin-bottom:.88rem;display:flex;align-items:center;gap:10px;background:#2D257D;padding:.56rem .9rem;border-radius:4px}.panel-title::before{content:'';width:3px;height:14px;background:var(--gold2);border-radius:2px;flex-shrink:0}.panel-note{padding:0 0 .72rem;color:var(--text3);font-size:10px;font-family:'DM Mono',monospace;letter-spacing:.05em;text-transform:uppercase}.table-wrap,.heatmap{overflow:auto;border:2px solid rgba(255,255,255,.22);border-radius:6px}
table{width:100%;border-collapse:collapse;font-size:12px}th{padding:8px 10px;text-align:right;font-size:10px;font-weight:800;letter-spacing:.08em;text-transform:uppercase;color:var(--text2);border-bottom:1px solid var(--border2);position:sticky;top:0;background:#111827}th:first-child,td:first-child{text-align:left}td{padding:8px 10px;text-align:right;border-bottom:1px solid rgba(255,255,255,.08);color:var(--text2);font-family:'DM Mono',monospace;font-size:11px;white-space:nowrap}td:first-child{color:#ffffff;font-family:'DM Sans',sans-serif;font-size:12px;font-weight:700}.row-pos{color:var(--green)!important}.row-neg{color:var(--red)!important}.wide-panel{margin-top:1px}
footer{border-top:2px solid var(--border2);padding:1rem 1.6rem;display:flex;align-items:center;justify-content:space-between}footer p{font-family:'DM Mono',monospace;font-size:10px;color:var(--text4);letter-spacing:.07em;text-transform:uppercase}
@media (max-width: 1400px){.chart-grid{grid-template-columns:repeat(2,minmax(0,1fr))!important}.insight-grid{grid-template-columns:1fr!important}}
@media (max-width: 980px){header{grid-template-columns:1fr;gap:10px;padding:12px 1rem}.logo,.header-center,.header-right{justify-self:center;align-items:center;text-align:center}.header-right{align-items:center}.toolbar,.page-bar,.view-bar,.active-filter{padding-left:1rem;padding-right:1rem}.kpi-strip{grid-template-columns:1fr!important}.chart-grid,.card-grid{grid-template-columns:1fr!important}.chart-card[data-span="half"],.chart-card[data-span="full"]{grid-column:auto}}
</style>
</head>
<body>
<header>
  <div class="logo">
    <div class="logo-mark"><svg viewBox="0 0 24 24"><polyline points="22 12 18 12 15 21 9 3 6 12 2 12"/></svg></div>
    <div>
      <div class="logo-title">FX PARITIES REPORT <span id="rangeInline" style="font-size:15px;color:var(--gold2);font-family:'DM Mono',monospace;margin-left:10px;letter-spacing:.03em"></span></div>
      <div class="logo-sub">EXECUTIVE DASHBOARD</div>
    </div>
  </div>
  <div class="header-center">
    <div class="header-center-title">FOREIGN EXCHANGE DASHBOARD</div>
    <div class="header-center-sub">TREASURY REPORTING TEAM</div>
  </div>
  <div class="header-right">
    <div class="header-stats"><div class="dyn-range" id="statsLine"></div><div class="clock" id="clock"></div></div>
    <div class="insight-box" id="headerInsightBox"><div class="insight-box-label">EXECUTIVE RADAR INSIGHT</div><div class="insight-box-text" id="headerInsightText"></div></div>
  </div>
</header>
<div class="toolbar">
  <div class="tool-chip"><label>QUICK RANGE</label><select id="quickRangeSelect"><option value="ALL">ALL</option><option value="1M">LAST 1M</option><option value="3M">LAST 3M</option><option value="6M">LAST 6M</option><option value="YTD">YTD</option></select><button class="clear-btn" id="clearQuick">×</button></div>
  <div class="tool-chip"><label>YEAR</label><select id="yearSelect"></select><button class="clear-btn" id="clearYear">×</button></div>
  <div class="tool-chip"><label>MONTH</label><select id="monthSelect"></select><button class="clear-btn" id="clearMonth">×</button></div>
  <div class="tool-chip"><label>FROM</label><input id="fromDate" type="date"></div>
  <div class="tool-chip"><label>TO</label><input id="toDate" type="date"></div>
  <div class="tool-chip"><label>DATES</label><button class="clear-btn" id="clearDates">×</button></div>
</div>
<div class="view-bar"><button class="view-btn on" data-view="DAILY">DAILY</button><button class="view-btn" data-view="WEEKLY">WEEKLY</button><button class="view-btn" data-view="MONTHLY">MONTHLY</button><button class="view-btn" data-view="YTD">YTD</button></div>
<div class="page-bar" id="pageBar"></div>
<div class="active-filter" id="activeFilterLabel"></div>
<div id="pageRoot" class="page-shell"></div>
<footer><p>FOREIGN EXCHANGE DASHBOARD · TREASURY REPORTING TEAM · UYGAR TALU</p><p id="footerRange"></p></footer>
<script>
const PAYLOAD = __PAYLOAD_JSON__;
const RECORDS = (PAYLOAD.records || []).map(r => ({...r})).sort((a,b) => (a.currency_code===b.currency_code ? a.date.localeCompare(b.date) : a.currency_code.localeCompare(b.currency_code)));
const META = PAYLOAD.meta || {};
const FOCUS = META.focus_currencies || [];
const OTHER_PAGE_SIZE = META.other_page_size || 10;
const COLORS = ["#5fb0ff","#ffb84d","#38d882","#ff6b6b","#c77dff","#41ead4","#ffd166","#7fd3ff","#ef476f","#06d6a0","#f78c6b","#c3f73a","#a29bfe","#00d2d3","#ff9f43","#70a1ff","#feca57"];
const state = { pageIndex:0, quickRange:'ALL', selectedYear:'ALL', selectedMonth:'ALL', fromDate:'', toDate:'', viewMode:'DAILY' };
const charts = {};
const allCurrencies = [...new Set(RECORDS.map(r => r.currency_code))].filter(c => c !== 'USD').sort();
const focusExisting = FOCUS.filter(c => allCurrencies.includes(c));
const otherCurrencies = allCurrencies.filter(c => !focusExisting.includes(c));
const years = [...new Set(RECORDS.map(r => String(r.year)))].sort();
const months = [...new Set(RECORDS.map(r => r.month_label))].sort();
const latestGlobalDate = (RECORDS[RECORDS.length-1] || {}).date || META.end_date;
function colorFor(code){ const idx = Math.abs([...code].reduce((s,ch)=>s+ch.charCodeAt(0),0)) % COLORS.length; return COLORS[idx]; }
function alpha(hex,a){ const h=hex.replace('#',''); const full=h.length===3 ? h.split('').map(x=>x+x).join('') : h; const int=parseInt(full,16); return `rgba(${(int>>16)&255},${(int>>8)&255},${int&255},${a})`; }
function fmtRate(v){ if(v===null || v===undefined || Number.isNaN(Number(v))) return '—'; return Number(v).toLocaleString('en-US',{minimumFractionDigits:4, maximumFractionDigits:6}); }
function fmtPct(v){ if(v===null || v===undefined || Number.isNaN(Number(v))) return '—'; return (Number(v)*100).toFixed(2)+'%'; }
function fmtPctSigned(v){ if(v===null || v===undefined || Number.isNaN(Number(v))) return '—'; const n=Number(v)*100; return (n>=0?'+':'')+n.toFixed(2)+'%'; }
function monthPretty(label){ if(!label || label==='ALL') return 'ALL'; const [y,m]=label.split('-'); const d=new Date(Number(y),Number(m)-1,1); return d.toLocaleString('en-GB',{month:'short',year:'numeric'}).toUpperCase(); }
function dateMs(d){ return new Date(d+'T00:00:00').getTime(); }
function endOfWeekKey(d){ const dt=new Date(d+'T00:00:00'); const day=(dt.getDay()+6)%7; dt.setDate(dt.getDate()-day+6); return dt.toISOString().slice(0,10); }
function updateClock(){ const n=new Date(); document.getElementById('clock').textContent = n.toLocaleDateString('en-GB',{day:'2-digit',month:'short',year:'numeric'}) + ' · ' + n.toLocaleTimeString('en-GB',{hour:'2-digit',minute:'2-digit'}); }
function buildPages(){ const defs=[]; const radarGroups=[]; if(focusExisting.includes('EUR')&&focusExisting.includes('GBP')) radarGroups.push(['EUR','GBP']); else { if(focusExisting.includes('EUR')) radarGroups.push(['EUR']); if(focusExisting.includes('GBP')) radarGroups.push(['GBP']); } if(focusExisting.includes('MYR')&&focusExisting.includes('SGD')) radarGroups.push(['MYR','SGD']); else { if(focusExisting.includes('MYR')) radarGroups.push(['MYR']); if(focusExisting.includes('SGD')) radarGroups.push(['SGD']); } const paired=new Set(radarGroups.flat()); focusExisting.forEach(c=>{ if(!paired.has(c)) radarGroups.push([c]); }); defs.push({label:'PAGE 1 · EXECUTIVE RADAR', title:'EXECUTIVE RADAR', currencies:focusExisting, chartGroups:radarGroups, type:'RADAR'}); let cursor=0,pageNo=2; while(cursor<otherCurrencies.length){ const chunk=otherCurrencies.slice(cursor,cursor+OTHER_PAGE_SIZE); defs.push({label:`PAGE ${pageNo}`, title:`ADDITIONAL CURRENCIES · PAGE ${pageNo}`, currencies:chunk, chartGroups:chunk.map(c=>[c]), type:'STANDARD'}); cursor+=OTHER_PAGE_SIZE; pageNo+=1; } return defs; }
const pages = buildPages();
function buildHeader(){ document.getElementById('rangeInline').textContent = `· ${META.start_date} – ${META.end_date}`; document.getElementById('statsLine').textContent = `${META.currencies_tracked} CURRENCIES · ${META.calendar_days} DAILY OBSERVATIONS`; document.getElementById('footerRange').textContent = `DATA RANGE: ${META.start_date} – ${META.end_date}`; }
function buildFilters(){ const yearSel=document.getElementById('yearSelect'); yearSel.innerHTML='<option value="ALL">ALL YEARS</option>'+years.map(y=>`<option value="${y}">${y}</option>`).join(''); const monthSel=document.getElementById('monthSelect'); monthSel.innerHTML='<option value="ALL">ALL MONTHS</option>'+months.map(m=>`<option value="${m}">${monthPretty(m)}</option>`).join(''); yearSel.addEventListener('change',e=>{ state.selectedYear=e.target.value; if(state.selectedYear!=='ALL') state.quickRange='ALL'; renderCurrentPage(); }); monthSel.addEventListener('change',e=>{ state.selectedMonth=e.target.value; if(state.selectedMonth!=='ALL') state.quickRange='ALL'; renderCurrentPage(); }); document.getElementById('quickRangeSelect').addEventListener('change',e=>{ state.quickRange=e.target.value; renderCurrentPage(); }); document.getElementById('fromDate').addEventListener('change',e=>{ state.fromDate=e.target.value||''; if(state.fromDate) state.quickRange='ALL'; renderCurrentPage(); }); document.getElementById('toDate').addEventListener('change',e=>{ state.toDate=e.target.value||''; if(state.toDate) state.quickRange='ALL'; renderCurrentPage(); }); document.getElementById('clearQuick').addEventListener('click',()=>{ state.quickRange='ALL'; document.getElementById('quickRangeSelect').value='ALL'; renderCurrentPage(); }); document.getElementById('clearYear').addEventListener('click',()=>{ state.selectedYear='ALL'; document.getElementById('yearSelect').value='ALL'; renderCurrentPage(); }); document.getElementById('clearMonth').addEventListener('click',()=>{ state.selectedMonth='ALL'; document.getElementById('monthSelect').value='ALL'; renderCurrentPage(); }); document.getElementById('clearDates').addEventListener('click',()=>{ state.fromDate=''; state.toDate=''; document.getElementById('fromDate').value=''; document.getElementById('toDate').value=''; renderCurrentPage(); }); document.querySelectorAll('.view-btn').forEach(btn=>btn.addEventListener('click',()=>{ state.viewMode=btn.dataset.view; document.querySelectorAll('.view-btn').forEach(b=>b.classList.toggle('on', b===btn)); renderCurrentPage(); })); }
function renderPageButtons(){ const bar=document.getElementById('pageBar'); bar.innerHTML=pages.map((p,i)=>`<button class="page-btn ${i===state.pageIndex?'on':''}" data-idx="${i}">${p.label}</button>`).join(''); bar.querySelectorAll('.page-btn').forEach(btn=>btn.addEventListener('click',()=>{ state.pageIndex=Number(btn.dataset.idx); renderCurrentPage(); })); }
function getFilteredBaseRows(currencies){ let rows=RECORDS.filter(r=>currencies.includes(r.currency_code)); if(state.selectedYear!=='ALL') rows=rows.filter(r=>String(r.year)===state.selectedYear); if(state.selectedMonth!=='ALL') rows=rows.filter(r=>r.month_label===state.selectedMonth); if(state.fromDate) rows=rows.filter(r=>r.date>=state.fromDate); if(state.toDate) rows=rows.filter(r=>r.date<=state.toDate); if(state.quickRange!=='ALL'&&state.selectedYear==='ALL'&&state.selectedMonth==='ALL'&&!state.fromDate&&!state.toDate){ const end=new Date(latestGlobalDate+'T00:00:00'); let start=new Date(end); if(state.quickRange==='1M') start.setDate(start.getDate()-30); if(state.quickRange==='3M') start.setDate(start.getDate()-92); if(state.quickRange==='6M') start.setDate(start.getDate()-183); if(state.quickRange==='YTD') start=new Date(end.getFullYear(),0,1); const startStr=start.toISOString().slice(0,10); rows=rows.filter(r=>r.date>=startStr&&r.date<=latestGlobalDate); } return rows; }
function aggregateRowsForView(rows){ if(state.viewMode==='DAILY') return rows.slice(); if(state.viewMode==='YTD'){ const last=rows.length ? rows[rows.length-1].date : latestGlobalDate; const year=String(last).slice(0,4); return rows.filter(r=>String(r.date).slice(0,4)===year); } const grouped={}; rows.forEach(r=>{ let key=r.date; if(state.viewMode==='WEEKLY') key=`${r.currency_code}|${endOfWeekKey(r.date)}`; if(state.viewMode==='MONTHLY') key=`${r.currency_code}|${r.month_label}`; if(!grouped[key]||r.date>grouped[key].date) grouped[key]=r; }); return Object.values(grouped).sort((a,b)=> a.currency_code===b.currency_code ? a.date.localeCompare(b.date) : a.currency_code.localeCompare(b.currency_code)); }
function latestRows(rows,currencies){ const map={}; rows.forEach(r=>{ if(!map[r.currency_code]||r.date>map[r.currency_code].date) map[r.currency_code]=r; }); return currencies.map(c=>map[c]).filter(Boolean); }
function chooseCols(count){ if(count<=1) return 1; if(count<=4) return 2; if(count<=9) return 3; return 4; }
function chartSpanMeta(count, cols, idx){ const remainder=count%cols; if(remainder===1&&idx===count-1&&count>1) return 'full'; if(remainder===2&&cols===4&&idx>=count-2) return 'half'; return ''; }
function destroyCharts(){ Object.keys(charts).forEach(k=>{ try{ charts[k].destroy(); }catch(e){} delete charts[k]; }); }
function activeFilterText(pageRows){ const bits=[]; if(state.quickRange!=='ALL') bits.push(`QUICK RANGE: ${state.quickRange}`); if(state.selectedYear!=='ALL') bits.push(`YEAR: ${state.selectedYear}`); if(state.selectedMonth!=='ALL') bits.push(`MONTH: ${monthPretty(state.selectedMonth)}`); if(state.fromDate||state.toDate) bits.push(`DATES: ${state.fromDate||META.start_date} TO ${state.toDate||latestGlobalDate}`); if(!bits.length) bits.push('FILTER: ALL AVAILABLE DATES'); if(pageRows.length) bits.push(`LATEST VISIBLE DATE: ${pageRows[pageRows.length-1].date}`); bits.push(`VIEW: ${state.viewMode}`); return bits.join(' · '); }
function buildExecutiveInsight(){ const radarRows=getFilteredBaseRows(focusExisting).sort((a,b)=>a.date.localeCompare(b.date)); const latestRadar=latestRows(radarRows, focusExisting); if(!latestRadar.length){ document.getElementById('headerInsightText').textContent='NO EXECUTIVE RADAR DATA IS AVAILABLE UNDER THE CURRENT FILTERS.'; return; } const valid=latestRadar.filter(r=>r.change_vs_prior_month_avg_pct!==null&&r.change_vs_prior_month_avg_pct!==undefined); const best=valid.slice().sort((a,b)=>Number(b.change_vs_prior_month_avg_pct)-Number(a.change_vs_prior_month_avg_pct))[0]; const worst=valid.slice().sort((a,b)=>Number(a.change_vs_prior_month_avg_pct)-Number(b.change_vs_prior_month_avg_pct))[0]; const vol=latestRadar.filter(r=>r.rolling_30d_volatility_pct!==null&&r.rolling_30d_volatility_pct!==undefined).slice().sort((a,b)=>Number(b.rolling_30d_volatility_pct)-Number(a.rolling_30d_volatility_pct))[0]; const positive=valid.filter(r=>Number(r.change_vs_prior_month_avg_pct)>=0).length; const total=valid.length; const dateLabel=latestRadar.slice().sort((a,b)=>a.date.localeCompare(b.date)).slice(-1)[0].date; document.getElementById('headerInsightText').innerHTML=`AS OF <strong>${dateLabel}</strong>, <strong>${positive}/${total}</strong> EXECUTIVE RADAR CURRENCIES ARE TRADING ABOVE THEIR PRIOR-MONTH AVERAGE. STRONGEST RELATIVE MOVE: <strong>${best ? best.currency_code + ' ' + fmtPctSigned(best.change_vs_prior_month_avg_pct) : '—'}</strong>. WEAKEST: <strong>${worst ? worst.currency_code + ' ' + fmtPctSigned(worst.change_vs_prior_month_avg_pct) : '—'}</strong>. HIGHEST 30D VOLATILITY: <strong>${vol ? vol.currency_code + ' ' + fmtPct(vol.rolling_30d_volatility_pct) : '—'}</strong>.`; }
function pointIndices(data){ const pts=data.map((v,i)=>({v:Number(v),i})).filter(x=>x.v===x.v&&x.v!==null&&x.v!==undefined); if(!pts.length) return null; let max=pts[0], min=pts[0], last=pts[pts.length-1]; pts.forEach(p=>{ if(p.v>max.v) max=p; if(p.v<min.v) min=p; }); return {max,min,last}; }
const annotatePlugin={ id:'annotateExtrema', afterDatasetsDraw(chart){ const ctx=chart.ctx; chart.data.datasets.forEach((ds,di)=>{ const meta=chart.getDatasetMeta(di); if(meta.hidden) return; const info=pointIndices(ds.data); if(!info) return; [{kind:'HIGH',idx:info.max.i,val:info.max.v,color:ds.borderColor},{kind:'LOW',idx:info.min.i,val:info.min.v,color:ds.borderColor},{kind:'LAST',idx:info.last.i,val:info.last.v,color:'#ffffff'}].forEach((it,n)=>{ const point=meta.data[it.idx]; if(!point) return; const props=point.getProps(['x','y'],true); const yShift=it.kind==='LOW' ? 16 : -14 - (n*2); ctx.save(); ctx.beginPath(); ctx.arc(props.x,props.y,4.2,0,Math.PI*2); ctx.fillStyle=it.color; ctx.fill(); ctx.strokeStyle='#07080b'; ctx.lineWidth=1.4; ctx.stroke(); const label=`${it.kind} ${fmtRate(it.val)}`; ctx.font="700 10px 'DM Mono'"; const tw=ctx.measureText(label).width; const x=Math.min(chart.chartArea.right - tw/2 - 4, Math.max(chart.chartArea.left + tw/2 + 4, props.x)); const y=props.y+yShift; ctx.fillStyle='rgba(7,8,11,.92)'; ctx.fillRect(x-tw/2-4,y-8,tw+8,16); ctx.strokeStyle=alpha(typeof ds.borderColor==='string'?ds.borderColor:'#ffffff', .9); ctx.strokeRect(x-tw/2-4,y-8,tw+8,16); ctx.fillStyle=it.kind==='LAST' ? '#ffffff' : ds.borderColor; ctx.textAlign='center'; ctx.textBaseline='middle'; ctx.fillText(label,x,y); ctx.restore(); }); }); }};
function baseLineOptions(){ return { responsive:true, maintainAspectRatio:false, interaction:{mode:'index',intersect:false}, plugins:{ legend:{labels:{color:'#ffffff'}}, tooltip:{callbacks:{label:(ctx)=>`${ctx.dataset.label}: ${fmtRate(ctx.parsed.y)}`}} }, scales:{ x:{ticks:{color:'#d0ccc6',maxTicksLimit:8},grid:{display:false},border:{color:'rgba(255,255,255,0.18)'}}, y:{ticks:{color:'#d0ccc6',callback:(v)=>fmtRate(v)},grid:{color:'rgba(255,255,255,0.07)'},border:{color:'transparent'}} } }; }
function buildSeries(rows,codes){ const filtered=rows.filter(r=>codes.includes(r.currency_code)).sort((a,b)=>a.date.localeCompare(b.date)||a.currency_code.localeCompare(b.currency_code)); const labels=[...new Set(filtered.map(r=>r.date))]; const datasets=codes.map(code=>{ const color=colorFor(code); const sub=filtered.filter(r=>r.currency_code===code); const map={}; sub.forEach(r=>map[r.date]=Number(r.usd_rate)); return {label:code,data:labels.map(d=>map[d]??null),borderColor:color,backgroundColor:alpha(color,.20),fill:true,borderWidth:2.6,pointRadius:0,tension:.20,spanGaps:true}; }); return {labels,datasets}; }
function renderCurrentPage(){ renderPageButtons(); destroyCharts(); buildExecutiveInsight(); const page=pages[state.pageIndex]; const baseRows=getFilteredBaseRows(page.currencies).sort((a,b)=>a.date.localeCompare(b.date)||a.currency_code.localeCompare(b.currency_code)); const viewRows=aggregateRowsForView(baseRows).sort((a,b)=>a.date.localeCompare(b.date)||a.currency_code.localeCompare(b.currency_code)); const latestForPage=latestRows(baseRows.length ? baseRows : RECORDS.filter(r=>page.currencies.includes(r.currency_code)), page.currencies); document.getElementById('activeFilterLabel').textContent=activeFilterText(baseRows); const root=document.getElementById('pageRoot'); if(!latestForPage.length){ root.innerHTML='<div class="section"><div class="section-title">NO DATA AVAILABLE</div><div class="section-note">THE CURRENT FILTERS RETURNED NO DATA FOR THIS PAGE.</div></div>'; return; } const best=latestForPage.filter(r=>r.change_vs_prior_month_avg_pct!==null&&r.change_vs_prior_month_avg_pct!==undefined).slice().sort((a,b)=>Number(b.change_vs_prior_month_avg_pct)-Number(a.change_vs_prior_month_avg_pct))[0]; const worst=latestForPage.filter(r=>r.change_vs_prior_month_avg_pct!==null&&r.change_vs_prior_month_avg_pct!==undefined).slice().sort((a,b)=>Number(a.change_vs_prior_month_avg_pct)-Number(b.change_vs_prior_month_avg_pct))[0]; const vol=latestForPage.filter(r=>r.rolling_30d_volatility_pct!==null&&r.rolling_30d_volatility_pct!==undefined).slice().sort((a,b)=>Number(b.rolling_30d_volatility_pct)-Number(a.rolling_30d_volatility_pct))[0]; root.innerHTML=`<div class="kpi-strip" id="kpiStrip"></div><div class="card-grid" id="cardGrid"></div><div class="section" style="padding:1.1rem 1.2rem .95rem"><div class="section-title">${page.title}</div><div class="section-note">${page.type==='RADAR' ? 'EXECUTIVE RADAR CURRENCIES WITH DIRECT MANAGEMENT FOCUS' : 'ADDITIONAL CURRENCIES AVAILABLE FOR DETAILED REVIEW'} · ${state.viewMode} VIEW</div></div><div class="chart-grid" id="chartGrid"></div><div class="insight-grid"><div class="panel"><div class="panel-title">LATEST PAGE SNAPSHOT</div><div class="panel-note">LATEST RATE · DOD · MTD · YTD · 30D VOLATILITY</div><div class="table-wrap" id="summaryTable"></div></div><div class="panel"><div class="panel-title">LATEST VS PRIOR MONTH AVG</div><div class="panel-note">TOP ABSOLUTE DEVIATIONS · LABELS DISPLAYED ON BARS</div><div class="chart-wrap" style="height:340px"><canvas id="moveBar"></canvas></div></div><div class="panel"><div class="panel-title">30D VOLATILITY RANKING</div><div class="panel-note">HIGHEST LATEST ROLLING VOLATILITY · LABELS DISPLAYED ON BARS</div><div class="chart-wrap" style="height:340px"><canvas id="volBar"></canvas></div></div></div><div class="panel wide-panel"><div class="panel-title">MONTH-END CHANGE HEATMAP</div><div class="panel-note">MONTHLY CHANGE MATRIX FOR ALL CURRENCIES DISPLAYED ON THE CURRENT PAGE</div><div class="heatmap" id="heatmapWrap"></div></div>`; document.getElementById('kpiStrip').innerHTML=[`<div class="kpi" style="--kc:var(--green)"><div class="kpi-label">STRONGEST VS PRIOR MONTH AVG</div><div class="kpi-val">${best ? best.currency_code + ' · ' + fmtPctSigned(best.change_vs_prior_month_avg_pct) : '—'}</div><div class="kpi-sub">${best ? best.currency_name + ' · LATEST RATE ' + fmtRate(best.usd_rate) : 'NO DATA AVAILABLE'}</div></div>`,`<div class="kpi" style="--kc:var(--red)"><div class="kpi-label">WEAKEST VS PRIOR MONTH AVG</div><div class="kpi-val">${worst ? worst.currency_code + ' · ' + fmtPctSigned(worst.change_vs_prior_month_avg_pct) : '—'}</div><div class="kpi-sub">${worst ? worst.currency_name + ' · LATEST RATE ' + fmtRate(worst.usd_rate) : 'NO DATA AVAILABLE'}</div></div>`,`<div class="kpi" style="--kc:var(--gold)"><div class="kpi-label">HIGHEST 30D VOLATILITY</div><div class="kpi-val">${vol ? vol.currency_code + ' · ' + fmtPct(vol.rolling_30d_volatility_pct) : '—'}</div><div class="kpi-sub">${vol ? vol.currency_name + ' · ROLLING STD OF DAILY MOVES' : 'NO DATA AVAILABLE'}</div></div>`].join(''); const latestMap={}; latestForPage.forEach(r=>latestMap[r.currency_code]=r); document.getElementById('cardGrid').innerHTML=page.currencies.map(code=>{ const r=latestMap[code]; if(!r) return `<div class="rate-card"><div class="rate-card-top"><div class="rate-card-code">${code}</div><div class="rate-card-name">NO DATA</div></div><div class="rate-card-rate">—</div><div class="rate-card-delta">NO ROWS UNDER CURRENT FILTERS</div></div>`; const delta=Number(r.change_vs_prior_month_avg_pct||0); const cls=delta>=0 ? 'pos' : 'neg'; return `<div class="rate-card"><div class="rate-card-top"><div class="rate-card-code">${r.currency_code}</div><div class="rate-card-name">${r.currency_name}</div></div><div class="rate-card-rate">${fmtRate(r.usd_rate)}</div><div class="rate-card-delta ${cls}">VS PRIOR MONTH AVG · ${fmtPctSigned(r.change_vs_prior_month_avg_pct)}</div></div>`; }).join(''); const chartGrid=document.getElementById('chartGrid'); const groups=page.chartGroups.filter(g=>g.some(code=>viewRows.some(r=>r.currency_code===code))); const cols=chooseCols(groups.length || 1); chartGrid.style.setProperty('--cols', cols); chartGrid.innerHTML=groups.map((group,idx)=>{ const span=chartSpanMeta(groups.length, cols, idx); const title=group.length===2 ? `${group[0]} + ${group[1]}` : `${group[0]} TREND`; const sub=group.length===2 ? 'AUTO-SCALED DUAL TRACKING' : 'FULL SERIES · SHADED AREA · HIGH LOW LAST'; return `<div class="chart-card" ${span ? `data-span="${span}"` : ''}><div class="chart-head"><div class="chart-title">${title}</div><div class="chart-sub">${sub}</div></div><div class="chart-wrap"><canvas id="chart_${idx}"></canvas></div></div>`; }).join(''); groups.forEach((group,idx)=>{ const {labels,datasets}=buildSeries(viewRows, group); charts[`chart_${idx}`]=new Chart(document.getElementById(`chart_${idx}`), {type:'line', data:{labels,datasets}, options: baseLineOptions(), plugins:[annotatePlugin]}); }); document.getElementById('summaryTable').innerHTML=`<table><thead><tr><th>CURRENCY</th><th>LATEST RATE</th><th>DOD</th><th>MTD</th><th>YTD</th><th>VS PRIOR MONTH AVG</th><th>30D VOL</th></tr></thead><tbody>${latestForPage.sort((a,b)=>a.currency_code.localeCompare(b.currency_code)).map(r=>`<tr><td>${r.currency_code}</td><td>${fmtRate(r.usd_rate)}</td><td class="${Number(r.rate_change_dod_pct||0)>=0?'row-pos':'row-neg'}">${fmtPctSigned(r.rate_change_dod_pct)}</td><td class="${Number(r.rate_change_mtd_pct||0)>=0?'row-pos':'row-neg'}">${fmtPctSigned(r.rate_change_mtd_pct)}</td><td class="${Number(r.rate_change_ytd_pct||0)>=0?'row-pos':'row-neg'}">${fmtPctSigned(r.rate_change_ytd_pct)}</td><td class="${Number(r.change_vs_prior_month_avg_pct||0)>=0?'row-pos':'row-neg'}">${fmtPctSigned(r.change_vs_prior_month_avg_pct)}</td><td>${fmtPct(r.rolling_30d_volatility_pct)}</td></tr>`).join('')}</tbody></table>`; const moveRows=latestForPage.filter(r=>r.change_vs_prior_month_avg_pct!==null&&r.change_vs_prior_month_avg_pct!==undefined).slice().sort((a,b)=>Math.abs(Number(b.change_vs_prior_month_avg_pct||0))-Math.abs(Number(a.change_vs_prior_month_avg_pct||0))).slice(0,Math.min(8,latestForPage.length)); charts.moveBar=new Chart(document.getElementById('moveBar'),{type:'bar',data:{labels:moveRows.map(r=>r.currency_code),datasets:[{data:moveRows.map(r=>Number(r.change_vs_prior_month_avg_pct||0)*100),backgroundColor:moveRows.map(r=>Number(r.change_vs_prior_month_avg_pct||0)>=0?alpha('#34d27c',.82):alpha('#ff6b6b',.82)),borderColor:moveRows.map(r=>Number(r.change_vs_prior_month_avg_pct||0)>=0?'#34d27c':'#ff6b6b'),borderWidth:1.4,borderRadius:6,barPercentage:.62,categoryPercentage:.72}]},options:{responsive:true,maintainAspectRatio:false,plugins:{legend:{display:false},tooltip:{callbacks:{label:(ctx)=>fmtPctSigned((ctx.parsed.y||0)/100)}}},scales:{x:{ticks:{color:'#d0ccc6',font:{family:'DM Mono',size:10,weight:'700'}},grid:{display:false},border:{color:'rgba(255,255,255,0.15)'}},y:{ticks:{color:'#d0ccc6',callback:(v)=>Number(v).toFixed(1)+'%',font:{family:'DM Mono',size:10,weight:'700'}},grid:{color:'rgba(255,255,255,0.07)'},border:{color:'transparent'}}}},plugins:[{afterDatasetsDraw(chart){ const ctx=chart.ctx; chart.getDatasetMeta(0).data.forEach((bar,i)=>{ const val=chart.data.datasets[0].data[i]; if(val===null||val===undefined) return; const p=bar.getProps(['x','y','base'],true); ctx.save(); ctx.font="700 10px 'DM Mono'"; ctx.fillStyle=val>=0 ? '#ffe3a0' : '#ffd1d1'; ctx.textAlign='center'; ctx.textBaseline='middle'; ctx.fillText((val>=0?'+':'') + Number(val).toFixed(2)+'%', p.x, val<0 ? p.base + 14 : p.y - 8); ctx.restore(); }); }}]}); const volRows=latestForPage.filter(r=>r.rolling_30d_volatility_pct!==null&&r.rolling_30d_volatility_pct!==undefined).slice().sort((a,b)=>Number(b.rolling_30d_volatility_pct||0)-Number(a.rolling_30d_volatility_pct||0)).slice(0,Math.min(8,latestForPage.length)); charts.volBar=new Chart(document.getElementById('volBar'),{type:'bar',data:{labels:volRows.map(r=>r.currency_code),datasets:[{data:volRows.map(r=>Number(r.rolling_30d_volatility_pct||0)*100),backgroundColor:volRows.map(r=>alpha(colorFor(r.currency_code),.84)),borderColor:volRows.map(r=>colorFor(r.currency_code)),borderWidth:1.4,borderRadius:6,barPercentage:.62,categoryPercentage:.72}]},options:{responsive:true,maintainAspectRatio:false,plugins:{legend:{display:false},tooltip:{callbacks:{label:(ctx)=>Number(ctx.parsed.y||0).toFixed(2)+'%'}}},scales:{x:{ticks:{color:'#d0ccc6',font:{family:'DM Mono',size:10,weight:'700'}},grid:{display:false},border:{color:'rgba(255,255,255,0.15)'}},y:{ticks:{color:'#d0ccc6',callback:(v)=>Number(v).toFixed(2)+'%',font:{family:'DM Mono',size:10,weight:'700'}},grid:{color:'rgba(255,255,255,0.07)'},border:{color:'transparent'}}}},plugins:[{afterDatasetsDraw(chart){ const ctx=chart.ctx; chart.getDatasetMeta(0).data.forEach((bar,i)=>{ const val=chart.data.datasets[0].data[i]; if(val===null||val===undefined) return; const p=bar.getProps(['x','y','base'],true); ctx.save(); ctx.font="700 10px 'DM Mono'"; ctx.fillStyle='#ffe3a0'; ctx.textAlign='center'; ctx.textBaseline='middle'; ctx.fillText(Number(val).toFixed(2)+'%', p.x, p.y - 8); ctx.restore(); }); }}]}); const monthSet=[...new Set(baseRows.map(r=>r.month_label))].sort(); const matrix={}; page.currencies.forEach(c=>matrix[c]={}); baseRows.forEach(r=>{ matrix[r.currency_code][r.month_label]=r.rate_change_mtd_pct; }); function heatBg(v){ if(v===null||v===undefined||Number.isNaN(Number(v))) return 'transparent'; const n=Number(v); if(n>0) return `rgba(52,210,124,${Math.min(.72,.16+Math.abs(n)*2.4)})`; if(n<0) return `rgba(255,107,107,${Math.min(.72,.16+Math.abs(n)*2.4)})`; return 'rgba(255,255,255,.06)'; } document.getElementById('heatmapWrap').innerHTML=`<table><thead><tr><th>CURRENCY</th>${monthSet.map(m=>`<th>${monthPretty(m)}</th>`).join('')}</tr></thead><tbody>${page.currencies.filter(c=>baseRows.some(r=>r.currency_code===c)).map(code=>`<tr><td>${code}</td>${monthSet.map(m=>`<td style="background:${heatBg(matrix[code][m])}">${fmtPct(matrix[code][m])}</td>`).join('')}</tr>`).join('')}</tbody></table>`; }
buildHeader(); buildFilters(); updateClock(); setInterval(updateClock, 1000); renderCurrentPage();
</script>
</body>
</html>'''
    return template.replace("__PAYLOAD_JSON__", json.dumps(payload, ensure_ascii=False))


def write_dashboard(path: Path, payload: dict) -> None:
    html = dashboard_html(payload)
    path.write_text(html, encoding="utf-8")


# =========================================================
# OPEN HELPERS
# =========================================================
def open_file(path: Path) -> None:
    try:
        system = platform.system().lower()
        if system == "windows":
            os.startfile(str(path))  # type: ignore[attr-defined]
        elif system == "darwin":
            subprocess.Popen(["open", str(path)])
        else:
            subprocess.Popen(["xdg-open", str(path)])
    except Exception as exc:  # noqa: BLE001
        warn(f"AUTO-OPEN FAILED FOR {path.name}: {exc}")


# =========================================================
# MAIN ORCHESTRATION
# =========================================================
@dataclass
class PreparedData:
    raw_df: pd.DataFrame
    master_df: pd.DataFrame
    currencies_df: pd.DataFrame
    start_date: date
    end_date: date
    source_mode: str


def prepare_data_via_api(today: date, load_ts: str) -> PreparedData:
    session = session_with_retries()
    currencies_df = fetch_active_currencies(session, today)
    currency_codes = currencies_df["iso_code"].astype(str).str.upper().tolist()
    observed = fetch_observed_rates_api(session, currency_codes, START_DATE, today, load_ts)
    raw_df = build_full_calendar_dataset(
        observed=observed,
        currencies_df=currencies_df,
        start_date=START_DATE,
        end_date=today,
        default_provider="FRANKFURTER",
        load_ts=load_ts,
    )
    master_df = build_masterdata(raw_df, currencies_df)
    return PreparedData(raw_df=raw_df, master_df=master_df, currencies_df=currencies_df, start_date=START_DATE, end_date=today, source_mode="API")


def prepare_data_via_upload(load_ts: str) -> PreparedData:
    file_path = ask_existing_file_path("PLEASE PROVIDE THE FILE PATH OF THE EXCEL FILE YOU HAVE")
    uploaded = load_user_excel_raw(file_path, load_ts)
    currencies_df = build_currency_catalog_from_upload(uploaded)
    start_date = min(uploaded["DATE"])
    end_date = max(uploaded["DATE"])
    raw_df = build_full_calendar_dataset(
        observed=uploaded.rename(columns={
            "DATE": "date",
            "CURRENCY_CODE": "currency_code",
            "USD_RATE": "usd_rate",
            "PROVIDER": "provider",
            "IS_FILLED": "is_filled",
            "LOAD_TIMESTAMP": "load_timestamp",
        }),
        currencies_df=currencies_df,
        start_date=start_date,
        end_date=end_date,
        default_provider="USER_UPLOAD",
        load_ts=load_ts,
    )
    master_df = build_masterdata(raw_df, currencies_df)
    return PreparedData(raw_df=raw_df, master_df=master_df, currencies_df=currencies_df, start_date=start_date, end_date=end_date, source_mode="USER_UPLOAD")


def choose_data_mode(today: date, load_ts: str) -> PreparedData:
    user_upload = ask_yes_no("DO YOU WANT TO GENERATE THE FX REPORT BY UPLOADING YOUR OWN DATA? (IF NO API CONNECTION WILL BE USED)")
    if user_upload:
        info("USER DATA MODE WAS SELECTED. THE FIRST SHEET OF THE PROVIDED EXCEL FILE WILL BE USED.")
        return prepare_data_via_upload(load_ts)

    info("API MODE WAS SELECTED. THE ENGINE WILL USE THE FX API CONNECTION.")
    return prepare_data_via_api(today, load_ts)


def main() -> None:
    today = local_today()
    load_ts = local_timestamp()
    print_banner(today)

    info("INITIALIZING THE FX PARITIES GENERATION PROCESS")
    prepared = choose_data_mode(today, load_ts)
    paths = build_output_paths(today)

    info("WRITING THE MAIN EXCEL WORKBOOK")
    write_workbook(paths["rates_xlsx"], prepared.raw_df, prepared.master_df)
    done(f"MAIN EXCEL WORKBOOK SAVED: {paths['rates_xlsx'].name}")

    info("WRITING THE STANDALONE MASTERDATA WORKBOOK")
    write_master_only_workbook(paths["master_xlsx"], prepared.master_df)
    done(f"MASTERDATA WORKBOOK SAVED: {paths['master_xlsx'].name}")

    info("BUILDING THE HTML DASHBOARD PAYLOAD")
    payload = build_dashboard_payload(prepared.master_df)

    info("WRITING THE HTML DASHBOARD")
    write_dashboard(paths["dashboard_html"], payload)
    done(f"HTML DASHBOARD SAVED: {paths['dashboard_html'].name}")

    info("OPENING ALL GENERATED OUTPUTS AUTOMATICALLY")
    open_file(paths["rates_xlsx"])
    open_file(paths["master_xlsx"])
    open_file(paths["dashboard_html"])

    finish_banner([paths["rates_xlsx"], paths["master_xlsx"], paths["dashboard_html"]])


if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        fail("THE PROCESS WAS CANCELLED BY THE USER.")
        sys.exit(1)
    except Exception as exc:  # noqa: BLE001
        fail(str(exc))
        input("PRESS ENTER TO CLOSE THE PROCESS")
        sys.exit(1)
